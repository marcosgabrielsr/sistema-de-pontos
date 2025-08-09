# Importando biblioteca tkinter e método adicional ttk
from tkinter import *

# Importando biblioteca para trabalhar com a planilha do excel
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import column_index_from_string

# Importando biblioteca para trabalhar com datas
from datetime import datetime, timedelta

# Importando pacote para gerar trabalhar com pastas
from func_pasta import Folders

# Importando funções e métodos da janela de colaboradores
from view_colab import *

# Importando funções e métodos da janela de pontos
from view_pontos import *

class Func_planilhas(Folders, Func_colab_bd, Func_bd_pontos):
    def __init__(self):
        self.colunas = ('D', 'E', 'F', 'G')

        self.col_inicio = column_index_from_string(self.colunas[0])
        self.col_fim = column_index_from_string(self.colunas[3])

        self.estilos_celulas = {
            'alinhamento_centro': Alignment(horizontal='center', vertical='center'),
            'preenchimento_falta': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),
            'preenchimento_atestado': PatternFill(start_color='ADCBFF', end_color='ADCBFF', fill_type='solid'),
            'fonte_falta': Font(color='FF0000', bold=True),
            'fonte_atestado': Font(color='0000FF', bold=True)
        }

    # Função que calcula dados de um mes referente a um ano
    def calcula_dados_mes(self, mes, ano):
        # Usa datetime para lidar com datas
        self.first_day_of_month = datetime(ano, mes, 1)
        self.last_day_of_month = (self.first_day_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        self.t_dias = (self.last_day_of_month - self.first_day_of_month).days + 1

    # Função que gera os arquivos excel de um colaborar
    def criar_planilha(self, mes, ano, nome_colab, faltas=True, lista_faltas=None):
        name_folder = f'{mes:02d}-{ano}'
        self.create_folder_mes(name_folder)
        self.calcula_dados_mes(mes, ano)

        # Carregando a planilha
        workbook = openpyxl.load_workbook("controle_de_ponto.xlsx")
        sheets = workbook.sheetnames

        sheet_config = workbook[sheets[0]]
        sheet_pontos = workbook[sheets[1]]

        sheet_config['C4'].value = self.first_day_of_month
        sheet_pontos['D4'].value = nome_colab
        
        cod = self.get_cod(nome_colab)[0][0]

        # Percorre a tabela dos pontos e preenche as células de D a G
        # partindo da linha 7 até à linha final determinada pelo total de dias do mês mais 7
        for i in range(7, self.t_dias + 7):
            dia = i - 6
            
            for col in self.colunas:
                # Algoritmo que faz a busca e pega o valor do horário referente ao dia desejado    
                self.conecta_pontos_bd()
                # Armazena a data de hoje
                data = datetime(ano, mes, dia)
                data = data.strftime("%d/%m/%Y")
                tipo = self.colunas.index(col)

                self.cursor.execute("""
                    SELECT horario FROM pontos
                    WHERE nome_pontos LIKE '%s' AND data LIKE '%s' AND tipo LIKE '%d'
                """ % (nome_colab, data, tipo)
                )

                busca = self.cursor.fetchall()
                self.desconecta_pontos_bd()

                # Pega os dados da busca caso eles existam
                sheet_pontos[f'{col}{i}'] = int(busca[0][0]) if (len(busca) > 0) else 0

            if faltas == True:
                # Verificação se os valores dos pontos estão todos vazios da linha
                valores = [sheet_pontos[f'{col}{i}'].value for col in self.colunas]

                if all(v == 0 for v in valores) and datetime(ano, mes, dia).weekday() != 6:
                    # Mesclando as células
                    sheet_pontos.merge_cells(start_row=i, start_column=self.col_inicio, end_row=i, end_column=self.col_fim)

                    celula = sheet_pontos[f'{self.colunas[0]}{i}']
                    celula.alignment = self.estilos_celulas['alinhamento_centro']
                    celula.value = 'Falta sem Justificativa'
                    celula.fill = self.estilos_celulas['preenchimento_falta']
                    celula.font = self.estilos_celulas['fonte_falta']
            else:
                # Verifica se a lista de dias de falta é diferente de None
                if lista_faltas is not None:
                    print(f'Lista de faltas: {lista_faltas}')
                    # Caso este dia esteja na lista de faltas, uma falta é definida na planilha
                    if i in lista_faltas:
                        # Mesclando as células
                        sheet_pontos.merge_cells(start_row=i, start_column=col_inicio, end_row=i, end_column=col_fim)

                        celula = sheet_pontos[f'{self.colunas[0]}{i}']
                        celula.alignment = self.estilos_celulas['alinhamento_centro']
                        celula.value = 'Falta sem Justificativa'
                        celula.fill = self.estilos_celulas['preenchimento_falta']
                        celula.font = self.estilos_celulas['fonte_falta']

        # definindo caminho para salvar a planilha
        path_new_sheet = f'planilhas/{mes:02d}-{ano}/{cod}-{mes:02d}-{ano}.xlsx'
        workbook.save(path_new_sheet)

    # Função que gera planilhas para uma lista de colaboradores
    def gerar_planilhas(self, mes, ano, f=True):
        for nome in self.get_colabs():
            self.criar_planilha(mes, ano, nome, faltas=f)

    # Função que insere um atestado à uma planilha
    def inserir_atestado(self, dia, path):
        # Abrindo planilha
        workbook = openpyxl.load_workbook(path)
        sheet_pontos = workbook[workbook.sheetnames[1]]
        # Acessando célula
        celula = sheet_pontos[f'{self.colunas[0]}{dia + 6}']
        print(f'{celula.value}')
        if celula.value == 'Falta sem Justificativa':
            celula.value = 'Falta com Atestado'
            celula.fill = self.estilos_celulas['preenchimento_atestado']
            celula.font = self.estilos_celulas['fonte_atestado']
            workbook.save(path)
            return 1
        elif celula.value == 'Falta com Atestado':
            return 2
        else:
            return 0
    
    # Função que insere um atestado à uma planilha
    def inserir_falta(self, dia, path):
        # Abrindo planilha
        workbook = openpyxl.load_workbook(path)
        sheet_pontos = workbook[workbook.sheetnames[1]]
        i = dia + 6
        # Mesclando as células
        sheet_pontos.merge_cells(start_row=i, start_column=self.col_inicio, end_row=i, end_column=self.col_fim)

        celula = sheet_pontos[f'{self.colunas[0]}{i}']
        celula.alignment = self.estilos_celulas['alinhamento_centro']
        celula.value = 'Falta sem Justificativa'
        celula.fill = self.estilos_celulas['preenchimento_falta']
        celula.font = self.estilos_celulas['fonte_falta']

        workbook.save(path)

    # Função que remove uma ou todas as faltas de um funcionário
    def remover_falta(self, dia, mes, ano, colab_rem_falta, path):
        # Abrindo planilha
        workbook = openpyxl.load_workbook(path)
        sheet_pontos = workbook[workbook.sheetnames[1]]
        self.calcula_dados_mes(mes, ano)

        # Listas de eventos importantes no mês
        lista_faltas = []
        lista_atestados = []
        lista_feriados = []

        # Percorre a planilha e anota os dias de atestado, feriado e faltas (diferente do dia a ser removido)
        for i in range(7, self.t_dias + 7):
            d = i - 6
            celula = sheet_pontos[f'{self.colunas[0]}{i}']
            if d != dia:
                if celula.value == 'Falta sem Justificativa':
                    lista_faltas.append(d)
                elif celula.value == 'Falta com Atestado':
                    lista_atestados.append(d)
                elif celula.value == 'Feriado':
                    lista_feriados.append(d)
        # Salva a planilha
        workbook.save(path)
        print(f'lista_faltas: {lista_faltas}')

        # self.criar_planilha(mes, ano, colab_rem_falta, faltas=False, lista_faltas=lista_faltas)
        # Adiciona os atestados
        for a in lista_atestados:
            self.inserir_atestado(a, path)
